import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import logging
import time
import pyautogui
from tkinter import simpledialog

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def close_specific_page(browser, page_url):
    try:
        logging.info(f"检查是否打开了页面: {page_url}")
        if browser.current_url == page_url:
            browser.find_element(By.TAG_NAME, 'body').send_keys(Keys.CONTROL + 'w')
            logging.info(f"已关闭页面: {page_url}")
    except Exception as e:
        logging.error(f"关闭页面时出现错误: {e}")

def select_checkbox_with_text(browser, search_keyword):
    li_elements = browser.find_elements(By.XPATH, f"//li[contains(span, '{search_keyword}')]")
    for li in li_elements:
        checkbox = li.find_element(By.XPATH, ".//input[@type='checkbox']")
        if not checkbox.is_selected():
            checkbox.click()
            logging.info(f"已选中复选框：{search_keyword}")
        else:
            logging.info(f"复选框已经是勾选状态：{search_keyword}")

def scroll_to_element(browser, element):
    try:
        actions = ActionChains(browser)
        actions.move_to_element(element).perform()
        logging.info(f"成功滚动到元素: {element.text}")
    except Exception as e:
        logging.error(f"滚动到元素时出现错误：{e}")

def wait_for_element_to_appear(browser, by, value, timeout=10):
    return WebDriverWait(browser, timeout).until(EC.presence_of_element_located((by, value)))

def validate_firefox_profile_path(profile_path):
    if not profile_path:
        return False
    if not os.path.exists(profile_path):
        return False
    return True

def save_firefox_profile_path(profile_path):
    with open("firefox_profile.json", "w") as f:
        f.write(profile_path)

def load_firefox_profile_path():
    try:
        with open("firefox_profile.json", "r") as f:
            return f.read().strip()
    except FileNotFoundError:
        return None

def process_link(browser, link, excel_file):
    try:
        logging.info(f"处理链接: {link}")

        browser.get(link)
        close_specific_page(browser, "https://app.importify.net/dashboard/")

        for handle in browser.window_handles:
            browser.switch_to.window(handle)
            if link in browser.current_url:
                break
        time.sleep(2)

        add_btn_con = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="addBtnCon"]')))
        add_btn_con.click()
        logging.info("点击了按钮//*[@id='addBtnCon']")

        try:
            element = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.XPATH, '//span[@class="inactive" and text()="Draft"]'))
            )
            logging.info("成功加载 Draft 元素")
            actions = ActionChains(browser)
            actions.move_to_element(element).perform()
            element.click()
            logging.info("成功点击 Draft 元素")
            time.sleep(2)

        except Exception as e:
            logging.error(f"等待和点击 Draft 元素时出现错误：{e}")

        select_button = WebDriverWait(browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@class="ms-choice"]'))
        )
        select_button.click()

        dropdown = WebDriverWait(browser, 10).until(
            EC.visibility_of_element_located((By.CLASS_NAME, "ms-drop"))
        )

        wb = load_workbook(excel_file)
        first_sheet_name = wb.sheetnames[0] if wb.sheetnames else "默认工作表名称"
        logging.info(f"获取的第一个工作表名称: {first_sheet_name}")

        search_keyword = first_sheet_name
        logging.info(f"搜索关键词: {search_keyword}")

        search_input = dropdown.find_element(By.CSS_SELECTOR, ".ms-search input[type='text']")
        search_input.clear()
        search_input.send_keys(search_keyword)
        logging.info(f"输入关键词: {search_keyword}")

        WebDriverWait(browser, 10)

        select_checkbox_with_text(browser, search_keyword)

        time.sleep(3)  # 可以根据实际情况调整等待时间

        try:
            variants_button = browser.find_element(By.XPATH,
                                                   '//button[@data-actab-id="2" and @data-actab-group="0"]')
            variants_button.click()
            logging.info("点击了 Variants 按钮")
            time.sleep(2)  # 等待页面加载

            # 选择 Import all variants automatically 单选框
            all_variants_radio = browser.find_element(By.ID, 'all_variants')
            all_variants_radio.click()
            logging.info("选择 Import all variants automatically 单选框")

            time.sleep(2)  # 等待页面反应

            # 选择 Select which variants to include 单选框
            price_switch_radio = browser.find_element(By.ID, 'price_switch')
            price_switch_radio.click()
            logging.info("选择 Select which variants to include 单选框")

            time.sleep(2)  # 等待页面反应
        except Exception as e:
            logging.error(f"点击 Variants 按钮时出现错误：{e}")



        add_to_store_button = browser.find_element(By.ID, 'addBtnSec')
        scroll_to_element(browser, add_to_store_button)

        add_to_store_button.click()
        logging.info("成功点击 Add to your Store 按钮")


        logging.info("操作完成")

        wait_for_element_to_appear(browser, By.ID, 'importify-app-container')
        logging.info("页面加载完成")
        time.sleep(30)


    except Exception as e:
        logging.error(f"执行操作时出现错误：{e}")

def main():
    logging.info("开始执行主程序")
    root = tk.Tk()
    root.withdraw()

    messagebox.showinfo("提示", "请先选择要导入的Excel文件")
    excel_file = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel files", "*.xlsx *.xls")])

    if not excel_file:
        messagebox.showerror("错误", "未选择文件，程序退出。")
        logging.error("未选择文件，程序退出。")
        return

    logging.info(f"选择的Excel文件: {excel_file}")

    os.startfile(excel_file)

    browser = None
    try:
        excel_window_title = os.path.basename(excel_file)
        excel_window = pyautogui.getWindowsWithTitle(excel_window_title)
        if excel_window:
            excel_window = excel_window[0]
            excel_window.activate()
            logging.info("激活Excel窗口")
            screen_width, screen_height = pyautogui.size()
            left_window_width = screen_width // 2
            left_window_height = screen_height
            excel_window.resizeTo(left_window_width, left_window_height)
            logging.info(f"调整窗口大小: {left_window_width}x{left_window_height}")
            excel_window.moveTo(0, 0)
            logging.info(f"移动窗口位置: (0, 0)")

        root = tk.Tk()
        root.withdraw()

        saved_profile_path = load_firefox_profile_path()

        if saved_profile_path and validate_firefox_profile_path(saved_profile_path):
            firefox_profile_path = saved_profile_path
        else:
            firefox_profile_path = None

        while not firefox_profile_path:
            profile_path = simpledialog.askstring("输入", "请输入Firefox配置文件路径：")
            if validate_firefox_profile_path(profile_path):
                firefox_profile_path = profile_path
                save_firefox_profile_path(firefox_profile_path)
            else:
                print("输入的路径无效，请重新输入。")

        options = Options()
        options.profile = firefox_profile_path
        browser = webdriver.Firefox(options=options)

        # 读取链接列表并处理
        links = []
        row_number = 3  # 从第3行开始读取
        column_number = "B"
        wb = load_workbook(excel_file)
        ws = wb.active
        cell_value = ws[f"{column_number}{row_number}"].value

        while cell_value:
            link = ws[f"{column_number}{row_number}"].hyperlink.target if ws[f"{column_number}{row_number}"].hyperlink else None
            if link:
                links.append(link)
            row_number += 1
            cell_value = ws[f"{column_number}{row_number}"].value

        if not links:
            logging.warning("未从Excel中读取到链接，无法处理页面")
            return

        for link in links:
            process_link(browser, link, excel_file)

    except Exception as e:
        logging.error(f"执行操作时出现错误：{e}")
    finally:
        if browser:
            browser.quit()
            logging.info("浏览器已关闭")

if __name__ == "__main__":
    main()
